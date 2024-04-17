from django.shortcuts import render,redirect
from django.urls import reverse
from .forms import ProductoForm,ProveedorForm,FacturaForm,ListaProductosFacturaForm,VentaProductosFacturaForm,ListaProductoSinUnidadFacturaForm,ProductoSinUnidadForm
from .models import *
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.http import HttpResponse
from openpyxl import Workbook
#Cierre de secion 
from django.contrib.auth import logout
from django.shortcuts import redirect
#Obtener el ID secion 
def ID_Session(request):
    session_id = request.session.session_key
    # Hacer algo con el ID de sesión
    return render(request, 'inicio.html', {'session_id': session_id})

# Decorador para restringir el acceso a vistas solo para administradores
def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated and request.user.username == 'Admin1':
            return view_func(request, *args, **kwargs)
        elif request.user.is_authenticated and request.user.username == 'Vendedor':
            return redirect(reverse_lazy('PageVentasLink'))  # Redirige a la página de ventas si el usuario es un vendedor
        else:
            return redirect('login')  # Redirige a la página de inicio de sesión si no está autenticado como Admin
    return wrapper

# Decorador para restringir el acceso a vistas solo para vendedores
def vendor_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated and request.user.username == 'Vendedor':
            return view_func(request, *args, **kwargs)
        else:
            return redirect(reverse_lazy('PageVentasLink'))  # Redirige a la página de ventas si no está autenticado como Vendedor
    return wrapper

def logout_view(request):
    logout(request)
    response = redirect('Login.html')
    # Eliminar la cookie de sesión de tu aplicación
    response.delete_cookie('sessionid')
    return response

######  PAGINAS DE PRINCIPALES  ######
@login_required
def PageInicio(request):
    return render(request,"inicio.html")

@login_required
@admin_required
def PageProductos(request):
    productos = Producto.objects.all()
    form = ProductoForm()
    form2 = ProductoSinUnidadForm()
    if request.method == "POST":
        termino_busqueda = request.POST['termino_busqueda']
        if termino_busqueda:
            # Filtrar productos por nombre, código, stock, código de barras, categoría, marca y código de producto
            productos = Producto.objects.filter(
                nombre__icontains=termino_busqueda
            ) | Producto.objects.filter(
                codigo__icontains=termino_busqueda
            ) | Producto.objects.filter(
                stock__icontains=termino_busqueda
            ) | Producto.objects.filter(
                codigo_barras__icontains=termino_busqueda
            ) | Producto.objects.filter(
                categoria__nombre_categoria__icontains=termino_busqueda
            ) | Producto.objects.filter(
                marca__nombre_marca__icontains=termino_busqueda
            ) | Producto.objects.filter(
                codigo_producto__icontains=termino_busqueda
            )
            contex={
            'formulario':form,
            'busqueda':productos,
            'alerta':"ver todos"
            }
            return render(request,"Productos-templates/productos.html",contex)
    contex={
        'productos':productos,
        'formulario':form,
        'formulario2':form2
    }
    return render(request,"Productos-templates/productos.html",contex)

######  PAGINAS RELACIONADAS A PRODUCTOS  ######
def PageProductosCategorias(request):
    categorias = Categoria.objects.all()
    contex={
        'categorias':categorias,
    }
    return render(request,"Productos-templates/productos_categorias.html",contex)

def PageProductosMarcas(request):
    marcas = Marca.objects.all()
    contex={
        'marcas':marcas,
    }
    return render(request,"Productos-templates/productos_marcas.html",contex)

def PageProductosProveedores(request):
    proveedores = Proveedor.objects.all()
    form = ProveedorForm()
    contex={
        'proveedores':proveedores,
        'formulario':form,
    }
    return render(request,"Productos-templates/productos_proveedores.html",contex)

# crear
def crearProducto(request):
    producto = ProductoForm(request.POST,request.FILES)
    if producto.is_valid():
        producto.save()
        return redirect('PageProductosLink')

def crearCategoria(request):
    if request.method == 'POST':
        dato1 = request.POST.get('int-codigo')
        dato2 = request.POST.get('txt-nombre')
        Categoria.objects.create(codigo_categoria=dato1,nombre_categoria=dato2)
        return redirect('PageProductosCategoriasLink')
        

def crearMarca(request):
    if request.method == 'POST':
        dato1 = request.POST.get('int-codigo')
        dato2 = request.POST.get('txt-nombre')
        Marca.objects.create(codigo_marca=dato1,nombre_marca=dato2)
        return redirect('PageProductosMarcasLink')
        
def crearProveedor(request):
    proveedor = ProveedorForm(request.POST)
    if proveedor.is_valid():
        proveedor.save()
        return redirect('PageProductosProveedoresLink')
       

# editar
def editarProducto(request,codigo_producto):
    producto = Producto.objects.get(codigo=codigo_producto)
    formulario = ProductoForm(instance=producto)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST ,request.FILES,instance=producto) 
        if form.is_valid():
            form.save()
            return redirect('PageProductosLink') 
    context={
        'producto':producto,
        'formulario':formulario,
        }
    return render(request,'Productos-templates/productos-editar.html',context=context)

def editarCategoria(request,codigo_categoria):
    categoria = Categoria.objects.get(codigo=codigo_categoria)
    if request.method == 'POST':
        dato1 = request.POST.get('int-codigo')
        dato2 = request.POST.get('txt-nombre')
        Categoria.objects.filter(codigo=categoria.codigo).update(codigo_categoria=dato1,nombre_categoria=dato2)

        return redirect('PageProductosCategoriasLink')
    context={
        'categoria':categoria,
    }
    return render(request,'Productos-templates/producto_categorias-editar.html',context)

def editarMarca(request,codigo_marca):
    marca = Marca.objects.get(codigo=codigo_marca)
    if request.method == 'POST':
        dato1 = request.POST.get('int-codigo')
        dato2 = request.POST.get('txt-nombre')
        Marca.objects.filter(codigo=marca.codigo).update(codigo_marca=dato1,nombre_marca=dato2)

        return redirect('PageProductosMarcasLink')
    context={
        'marca':marca,
    }
    return render(request,'Productos-templates/producto_marcas-editar.html',context)

def editarProveedor(request,codigo_proveedor):
    proveedor = Proveedor.objects.get(codigo=codigo_proveedor)
    formulario = ProveedorForm(instance=proveedor)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST,instance=proveedor) 
        if form.is_valid():
            form.save()
            return redirect('PageProductosProveedoresLink') 
    context={
        'proveedor':proveedor,
        'formulario':formulario,
        }
    return render(request,'Productos-templates/producto_proveedor-editar.html',context=context)

# eliminar
def eliminarProducto(request,codigo_producto):
    Producto.objects.get(codigo=codigo_producto).delete()
    return redirect('PageProductosLink')

def eliminarCategoria(request,codigo_categoria):
    Categoria.objects.get(codigo=codigo_categoria).delete()
    return redirect('PageProductosCategoriasLink')
    
def eliminarMarca(request,codigo_marca):
    Marca.objects.get(codigo=codigo_marca).delete()
    return redirect('PageProductosMarcasLink')

def eliminarProveedor(request,codigo_proveedor):
    Proveedor.objects.get(codigo=codigo_proveedor).delete()
    return redirect('PageProductosProveedoresLink')

# detalles
def detalleProducto(request,codigo_producto):
    producto =  Producto.objects.get(codigo=codigo_producto)
    return render(request,'Productos-templates/producto_detalles.html',context={
        'producto':producto
    })


@login_required
def PageVentas(request): 
    if request.method == 'POST':
        factura = FacturaForm(request.POST)
        if factura.is_valid():
            # guardamos el codigo de la factura
            fac = factura.save()
            # enviamos el codigo a otra vista
            return redirect('crearVentaLink',fac.codigo)
        
    # Obtener todas las facturas que tienen ventas asociadas
    facturas_con_ventas = Factura.objects.filter(venta_productos_factura__isnull=False).distinct()
    formulario_fac = FacturaForm()
    context={ 
        'fact':facturas_con_ventas,
        'form_fac':formulario_fac,
    }
    return render(request,"Ventas-templates/venta.html",context)

def crearVenta(request,cod_fac):
    # todos los productos que tiene la misma factura
    lista_productos = Lista_Productos_Factura.objects.filter(codigo_factura=cod_fac)
    lista_productos_su = Lista_ProductoSinUnidad_Factura.objects.filter(codigo_factura=cod_fac)
    
    # subtottal lista productos unidad
    subtotalUNI = [ float(lp.producto.precio * lp.cantidad_vendida) for lp in lista_productos ]
    # subtottal lista productos gramos
    subtotalGRA = [ round(float(lp.cantidad_vendida_gramos)/ 1000 * float(lp.producto.precio_por_kilo) ,1) for lp in lista_productos_su ]
    
    # listas cobn sus subtotales
    prueba_produc_list_uni = zip(lista_productos,subtotalUNI)
    prueba_produc_list_gra = zip(lista_productos_su,subtotalGRA)
    

    # total de la venta
    total = sum(subtotalUNI + subtotalGRA)
    
    # formulario del listado de productos con un valor inicial de el codigo de fac 
    form_listado_productos = ListaProductosFacturaForm(initial={'codigo_factura': cod_fac})
    # formulario del listado de productos sin unidad con el valor del codigo de fac
    form_listado_productos_sin_unidad = ListaProductoSinUnidadFacturaForm(initial={'codigo_factura': cod_fac,})
    # formulario de venta
    formulario_venta = VentaProductosFacturaForm(initial={'factura': cod_fac,'total':total})
    #error de stock
    error_stock = ""
    if request.method == 'POST':
        if 'formulario1' in request.POST:
            form = ListaProductosFacturaForm(request.POST)
            #obtenemos el codigo del producto
            cod_producto =request.POST.get('producto')
            #obtenemos el producto
            prod = Producto.objects.get(codigo=cod_producto)
            #obtenemos el stock del producto
            stock_producto = prod.stock
            cantidad = request.POST.get('cantidad_vendida')   
            if stock_producto > int(cantidad):
                if form.is_valid():
                    form.save()
                    return redirect(request.path)
            else:
                error_stock = "no se puede agregar el producto por falta de stock"
        else :
            # factura instacia
            fac = Factura.objects.get(codigo=cod_fac)
            #obtenemos el codigo del producto
            cod_producto =request.POST.get('producto')
            #obtenemos el producto
            prod = ProductoSinUnidad.objects.get(codigo=cod_producto)
             #obtenemos el stock del producto en gramos 
            stock_producto_su = prod.stock_en_kilos * 1000
            # obtenemos el precio por kilo
            precio_kilo = prod.precio_por_kilo
              
            if 'cantidad_vendida_gramos' in request.POST: #se vende por gramos
                cant_gramos =request.POST.get('cantidad_vendida_gramos')
                if stock_producto_su > int(cant_gramos):

                    gramos = float(cant_gramos)
                    Lista_ProductoSinUnidad_Factura.objects.create(producto=prod,cantidad_vendida_gramos=gramos,codigo_factura=fac)
                
                    return redirect(request.path)
                else:
                    error_stock = "no se puede agregar el producto por falta de stock"
                
            else: #se vende por dinero
                can_dinero = request.POST.get('ventaDinero')
                
                total_en_gramos = float(can_dinero) / float(precio_kilo) * 1000.00  # Convertir dinero a gramos
                
                Lista_ProductoSinUnidad_Factura.objects.create(producto=prod,cantidad_vendida_gramos=total_en_gramos,codigo_factura=fac)

                return redirect(request.path)
                 
    context={
        # 'lista_productos_su':lista_productos_su,
        # 'lista_productos':lista_productos,
        # 'subtotalUNI':subtotalUNI,
        # 'subtotalGRA':subtotalGRA, 
        
        'form_listado_productos':form_listado_productos,
        'formulario_venta':formulario_venta,
        'error_stock':error_stock ,
        'form_listado_productos_sin_unidad':form_listado_productos_sin_unidad,
        
        'prueba_produc_list_uni':prueba_produc_list_uni,
        'prueba_produc_list_gra':prueba_produc_list_gra
        
    }
    return render(request,'Ventas-templates/crear-venta.html',context=context)

def procesarVenta(request,cod_fac):
    if request.method == 'POST':
        # obtenemos todos los productos vendidos
        lista_productos = Lista_Productos_Factura.objects.filter(codigo_factura=cod_fac)
        lista_productos_2 = Lista_ProductoSinUnidad_Factura.objects.filter(codigo_factura=cod_fac)
        # obtenemos todos los productos
        productos = Producto.objects.all()
        productos_2 = ProductoSinUnidad.objects.all()
         # quitar el stok en los productos por unidad
        for lisProd in lista_productos:
            for pro in productos:
                # preguntamos si el codigo del producto lista es el mismo que el de product
                if lisProd.producto.codigo == pro.codigo:
                    # obtenemos el codigo del producto y le restamos la catidad 
                    productoObtenido = Producto.objects.get(codigo=pro.codigo)
                    # validamos el stock
                    if productoObtenido.stock > lisProd.cantidad_vendida:
                       productoObtenido.stock -= lisProd.cantidad_vendida
                       productoObtenido.save()
                       
        for lpsu in lista_productos_2:
            for pro in productos_2:
                if lpsu.producto.codigo == pro.codigo:
                    productoObtenido = ProductoSinUnidad.objects.get(codigo=pro.codigo)
                    can_ven_kil = lpsu.cantidad_vendida_gramos / 1000
                    if productoObtenido.stock_en_kilos > can_ven_kil :
                       productoObtenido.stock_en_kilos -= can_ven_kil
                       productoObtenido.save()
                    
        # validacion de factura vacia   
        form = VentaProductosFacturaForm(request.POST)
        if lista_productos.exists() or lista_productos_2.exists():  #verifica si hay datos    
            if form.is_valid():
                form.save() 
                return redirect('detalleVentaLink',cod_fac)       
        else:
            # eliminamos la factura 
            Factura.objects.get(codigo=cod_fac).delete()
            # mandamos el error a una template nueva
            error = "error creaste una factura sin productos vuelve a crear una factura valida"
            return render(request,'Ventas-templates/error-venta.html',context={'error':error})
    return redirect('PageVentasLink')

def eliminarProductoLista(request,cod_prod_list):
    producto_lista = Lista_Productos_Factura.objects.get(codigo=cod_prod_list).delete()
    url_pagina_anterior = request.META.get('HTTP_REFERER')
    # Redireccionar a la página anterior
    return redirect(url_pagina_anterior)

def eliminarProductoListaSU(request,cod_prod_list_g):
    Lista_ProductoSinUnidad_Factura.objects.get(codigo=cod_prod_list_g).delete()
    url_pagina_anterior = request.META.get('HTTP_REFERER')
    # Redireccionar a la página anterior
    return redirect(url_pagina_anterior)


def eliminarVenta(request,cod_fac): 
    Factura.objects.get(codigo=cod_fac).delete()
    return redirect('PageVentasLink') 


def deleteVentaTemplate(request,cod_fac):
    venta = Venta_Productos_Factura.objects.get(factura=cod_fac)
    lista = Lista_Productos_Factura.objects.filter(codigo_factura=cod_fac)
    cadena = ""
    for objeto in lista:
        cadena += str(objeto) + ", "
    historial.objects.create(venta=venta.codigo,productos=cadena,metodo=venta.metodo,total=venta.total)
    
    Factura.objects.get(codigo=cod_fac).delete()
    return redirect('PageVentasLink')

def detalleVenta(request,cod_fac):
    factura =Factura.objects.get(codigo=cod_fac)
    
    lista_productos = Lista_Productos_Factura.objects.filter(codigo_factura=cod_fac)
    lista_productos_2 = Lista_ProductoSinUnidad_Factura.objects.filter(codigo_factura=cod_fac)
    
    venta = Venta_Productos_Factura.objects.get(factura=cod_fac)
    titulo = f"Tiket {factura}"
    lugar = "Puente Piedra"
    telefono = '52345432532'
    
    context = {
        'factura':factura,
        'lista_productos':lista_productos,
        'lista_productos_2':lista_productos_2,
        'venta':venta,
        'titulo':titulo,
        'lugar':lugar,
        'telefono':telefono,
    }
    return render(request,'Ventas-templates/detalle-venta.html',context)

from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

def generate_pdf(request, titulo, fecha, lugar, metodo_pago, codigo):
    
    # Crear el contenido del PDF
    pdf_content = []
    pdf_content.append(Paragraph(titulo, getSampleStyleSheet()['Title']))
    pdf_content.append(Paragraph(f"Fecha y Hora: {fecha}", getSampleStyleSheet()['BodyText']))
    pdf_content.append(Paragraph(f"Lugar: {lugar}", getSampleStyleSheet()['BodyText']))
    pdf_content.append(Paragraph(f"Método de Pago: {metodo_pago}", getSampleStyleSheet()['BodyText']))
    pdf_content.append(Paragraph(f"Código: {codigo}", getSampleStyleSheet()['BodyText']))

    lista = Lista_Productos_Factura.objects.filter(codigo_factura=codigo)
    lista_productos_2 = Lista_ProductoSinUnidad_Factura.objects.filter(codigo_factura=codigo)
    # venta = Venta_Productos_Factura.objects.get(factura=codigo)
    data = [["Producto", "Precio unitario", "Cantidad", "Total"]]
    
    subtotalUNI = [ float(lp.producto.precio * lp.cantidad_vendida) for lp in lista  ]
    subtotalGRA = [ round(float(lp.cantidad_vendida_gramos)/ 1000 * float(lp.producto.precio_por_kilo) ,1) for lp in lista_productos_2 ] 
    
    total = sum(subtotalUNI  + subtotalGRA)
    for lis in lista:
        subtotal = lis.producto.precio * lis.cantidad_vendida
        data.append([lis.producto, f"${lis.producto.precio}", str(lis.cantidad_vendida), f"${subtotal}"])
    for lis in lista_productos_2:
        subtotal = lis.producto.precio_por_kilo * (lis.cantidad_vendida_gramos / 1000)
        data.append([lis.producto, f"${lis.producto.precio_por_kilo}", str(lis.cantidad_vendida_gramos), f"${subtotal}"])
    
    
    
    
    
    table = Table(data)
    style = TableStyle([('TEXTCOLOR', (0, 0), (-1, -1), (0, 0, 0)),  # Color de texto: negro
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0))])  # Color de la línea de la tabla: negro
    table.setStyle(style)
    pdf_content.append(table)

    cantidad_productos = len(lista) if lista else len(lista_productos_2)
    pdf_content.append(Paragraph(f"Total de Productos: {cantidad_productos}", getSampleStyleSheet()['BodyText']))
    pdf_content.append(Paragraph(f"Total: ${total}", getSampleStyleSheet()['BodyText']))

    # Generar el PDF
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=letter)
    pdf.build(pdf_content)

    # Devolver el PDF como respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ticket_venta.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response

#ESTADISTICAS 
from json import dumps

@admin_required
def PageEstadisticas(request):
    p = Producto.objects.all()
    c = Categoria.objects.all()
    m = Marca.objects.all()
    pv = Proveedor.objects.all()
    ventas = Venta_Productos_Factura.objects.all()
    
    
    productos_nombres =  [prod.nombre  for prod in p]
    productos_stocks = [prod.stock  for prod in p]
    categorias_nombres = [cat.nombre_categoria for cat in c]
    categorias_cantidad = [1 for cat in c]
    marcas_nombres = [mar.nombre_marca for mar in m]
    marcas_cantidad = [1 for mar in m]
    proveedor_nombre = [prov.nombres  for prov in pv]
    proveedor_cantidad = [1  for prov in pv]
    venta_total_dinero = [v.total for v in ventas]
    venta_fecha = [str(v.factura.fecha_factura.strftime('%d/%m/%y')) for v in ventas]
    
    print(venta_fecha)
    
    context = {
    'productos_nombres': productos_nombres,
    'productos_stocks': productos_stocks,
    'categorias_nombres': categorias_nombres,
    'categorias_cantidad':categorias_cantidad,
    'marcas_nombres': marcas_nombres,
    'marcas_cantidad':marcas_cantidad,
    'proveedor_nombre': proveedor_nombre,
    'proveedor_cantidad':proveedor_cantidad,
    'venta_total_dinero': venta_total_dinero,
    'venta_fecha': venta_fecha,
}
    dataJSON = dumps(context) 
    return render(request,'estadisticas.html',{'data':dataJSON})


from datetime import datetime
from django.db.models import Sum
from calendar import monthrange

# REPORTES
@admin_required
def reportesPage(request):
    ventas = Venta_Productos_Factura.objects.all()
    ventas_totales = sum([ v.total  for v in ventas])
     # Obtén el año actual
    año_actual = datetime.now().year

    ventas_por_mes = []
    for mes in range(1, 13):  # Iterar sobre los 12 meses
        # Obtener el rango de días del mes actual
        inicio_mes = datetime(año_actual, mes, 1)
        fin_mes = datetime(año_actual, mes, monthrange(año_actual, mes)[1])

        # Filtrar las ventas por el mes actual
        ventas_mes_actual = Venta_Productos_Factura.objects.filter(
            factura__fecha_factura__range=(inicio_mes, fin_mes)
        ).aggregate(total_mes_actual=Sum('total'))['total_mes_actual']

        # Agregar el total de ventas del mes actual a la lista
        ventas_por_mes.append({
            'mes': mes,
            'total': ventas_mes_actual or 0  # Si no hay ventas, establecer el total en 0
        })
    
    ventas = [v['total'] for v in ventas_por_mes]
    
    
    
    context={
        'ventas_totales_por_mes':dumps(ventas),
        'ventas_totales':ventas_totales
    }
    return render(request,'reportes.html',context)

# Imprimir excel de productos
def descargar_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'

    # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()
    # Obtener la hoja de cálculo activa
    ws = wb.active

    # Agregar encabezados
    ws.append(['Código Producto', 'Código Barras', 'Nombre', 'Precio', 'Stock', 'Categoría', 'Marca', 'Proveedor'])

    # Agregar datos de productos
    productos = Producto.objects.all()
    for producto in productos:
        ws.append([producto.codigo_producto, producto.codigo_barras, producto.nombre, producto.precio, producto.stock, producto.categoria.nombre_categoria, producto.marca.nombre_marca, producto.proveedor.nombres])

    # Guardar el libro de trabajo
    wb.save(response)

    return response

# Imprimir excel de categoria
def descargar_excel_categoria(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Categorias.xlsx"'

    # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()
    # Obtener la hoja de cálculo activa
    ws = wb.active

    # Agregar encabezados
    ws.append(['Código', 'Nombre'])

    # Agregar datos de categorías
    categorias = Categoria.objects.all()
    for categoria in categorias:
        ws.append([categoria.codigo_categoria, categoria.nombre_categoria])

    # Guardar el libro de trabajo
    wb.save(response)

    return response

# Imprimir excel de Marcas
def descargar_excel_marca(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Marca.xlsx"'

    # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()
    # Obtener la hoja de cálculo activa
    ws = wb.active

    # Agregar encabezados
    ws.append(['Código', 'Nombre'])

    # Agregar datos de marcas
    marcas = Marca.objects.all()
    for marca in marcas:
        ws.append([marca.codigo_marca, marca.nombre_marca])

    # Guardar el libro de trabajo
    wb.save(response)

    return response

# Imprimir excel de Proveedores
def descargar_excel_proveedores(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Proveedores.xlsx"'

    # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()
    # Obtener la hoja de cálculo activa
    ws = wb.active

    # Agregar encabezados
    ws.append(['Código de Proveedor', 'Nombres', 'Apellidos', 'RUC/DNI', 'Dirección', 'Correo Electrónico', 'Número de Teléfono'])

    # Agregar datos de proveedores
    proveedores = Proveedor.objects.all()
    for proveedor in proveedores:
        ws.append([proveedor.codigo_proveedor, proveedor.nombres, proveedor.apellidos, proveedor.ruc_dni, proveedor.direccion, proveedor.correo_electronico, proveedor.numero_telefono])

    # Guardar el libro de trabajo
    wb.save(response)

    return response


# historial
@admin_required
def historialPage(request):
    histo = historial.objects.all()
    ventas = Venta_Productos_Factura.objects.all()
    context={
        'historial_eliminados':histo,
        'ventas_template':ventas,
    }
    return render(request,'historial.html',context)